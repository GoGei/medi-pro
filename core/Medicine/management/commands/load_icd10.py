import io
import zipfile
from openpyxl import load_workbook
from django.db.transaction import atomic
from django.core.management.base import BaseCommand, CommandError

from core.Medicine.enums import MedicineHandbookSources
from core.Medicine.models import ICD10


class Command(BaseCommand):
    """
    https://www.cms.gov/medicare/coordination-benefits-recovery/overview/icd-code-lists
    """
    help = 'Load ICD-10 codes from HL7 FHIR'

    @atomic
    def handle(self, *args, **options):
        archive = zipfile.ZipFile('core/Medicine/fixture/icd-10.zip', 'r')
        table_data = archive.read('section111validicd10-jan2025_0.xlsx')
        wb = load_workbook(io.BytesIO(table_data), read_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        headers = next(rows)
        created_count = updated_count = total = archived = 0

        for row in rows:
            item = dict(zip(headers, row))
            total += 1

            try:
                icd10, created = ICD10.objects.update_or_create(
                    code=item['CODE'],
                    defaults={
                        'name': item['SHORT DESCRIPTION (VALID ICD-10 FY2025)'],
                        'source': MedicineHandbookSources.CMS_GOV
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

                if item['NF EXCL'] == 'Y':
                    icd10.archive()
                    archived += 1

                if total % 500 == 0:
                    self.stdout.write(f'Processed {total} ICD-10 codes')
            except Exception as e:
                raise CommandError(f'For item {item} raised error: {e}')

        self.stdout.write(self.style.SUCCESS(f'Successfully created {created_count} ICD-10 codes'))
        self.stdout.write(self.style.SUCCESS(f'Successfully updated {updated_count} ICD-10 codes'))
        self.stdout.write(self.style.WARNING(f'Archived: {archived} ICD-10 codes'))
